import os
import time as t
import pyautogui as pe
import webbrowser as webb
import pyperclip as pc
from openpyxl import Workbook, load_workbook

#=====================================================#
#===================Functions=========================#
#=====================================================#

def GetTime():

    global day #global - declare varibales to use in all the code
    global month_day
    global hour
    global sec
    global minute
    global month
    global year
    
    current = t.gmtime() #getting the time into veribales
    day = current[6]
    month_day = current[2]
    
#day == 1:#thusday
       
#day == 2:#wensday
       
#day == 3:#thursday
       
#day == 4:#friday
   
#day == 5:#saturday
       
#day == 6:#sunday
       
#day == 0:#monday
 
    month = current[1]
    year =  current[0]

    hour = current[3]+2
    if hour == 23+2:
        hour = 1
    if hour == 22+2:
        hour = 0

    minute = current[4]
    sec = current[5]

    t.sleep(1)
#-----------------------------#
def GetXlVal(): #getting info from the xl file
    
    global course_name_1
    global course_name_2
    global zoom_link
    global lecturer_name
    global wb
    global ws
    global i

    #course_name_1 = ws.cell(row=, column=) #need to copy every time

    for i in range(2,8): #find the lecterer name and zoom liink
        course_name_2 = ws.cell(row=20, column=i)
        
        if course_name_1.value == course_name_2.value: #getting info

            zoom_link = ws.cell(row=21, column=i)
            lecturer_name = ws.cell(row=22, column=i)
            break
#-----------------------------#
def Filmora(): #open filmora

    #GetTime()
    #print('Open Filmora',hour,':', minute,':', sec)
    
    os.startfile('C:/Albert/Programs/Filmora/Filmora.exe')
    t.sleep(15)

    cords = pe.locateCenterOnScreen("Media//Filmora//OpenFilmora.png", confidence=0.8)
    pe.click(cords)
    t.sleep(15)

    #open the record section
    cords = pe.locateCenterOnScreen("Media//Filmora//FIleFilmora.jpg", confidence=0.8) 
    pe.click(cords)
    t.sleep(1)
    pe.press('down', presses=9)
    t.sleep(1)
    pe.press('right')
    t.sleep(1)
    pe.press('down')
    t.sleep(1)
    pe.press('enter')
    t.sleep(5)

    ##next step in the recording - desktop
    #cords = pe.locateCenterOnScreen("Media//Filmora//‏‏SettingsFilmora.png", confidence=0.9)
    #pe.click(cords)
    #t.sleep(1)
    #cords = pe.locateCenterOnScreen("Media//Filmora//SaveLocationFilmora.jpg", confidence=0.8)
    #pe.click(cords)
    #t.sleep(1)
    #cords = pe.locateCenterOnScreen("Media//Filmora//DesktopFilmora.jpg", confidence=0.8)
    #pe.click(cords)
    #t.sleep(1)
    #pe.press('enter')
    #t.sleep(1)
    #cords = 0
    
    #No mic
    cords = pe.locateCenterOnScreen("Media//Filmora//NoMicFilmora.png")
    if cords != 0:
        pe.click(cords)

    t.sleep(15)
#--------------------------------#
def StartRecFilmora():

    GetTime()
    
    cords = pe.locateCenterOnScreen("Media//Filmora//StartRecFilmora.jpg", confidence=0.8)
    print('Lets start Recording',hour,':', minute,':', sec)
    pe.click(cords)    
    t.sleep(10)
#--------------------------------#
def end_record():

    print('Lecture Done')
    pe.press('f9')
    GetTime()
    print('Recording Ended',hour,':', minute)
    t.sleep(10)
#--------------------------------#
def close_Filmora():

    cords = pe.locateCenterOnScreen("Media//Filmora//CloseFilmora.png", confidence=0.9)
    pe.click(cords)
    t.sleep(5)

    with pe.hold('alt'):
        t.sleep(1)
        pe.press('f4')
        t.sleep(1)

    cords = pe.locateCenterOnScreen("Media//Filmora//CloseRecFilmora.png", confidence=0.9)
    pe.click(cords)
    t.sleep(5)

    with pe.hold('alt'):
        t.sleep(1)
        pe.press('f4')
        t.sleep(1)
    
    GetTime()
    print('Filmora Closed',hour,':', minute,':', sec)
    t.sleep(5)
#--------------------------------#
def zoom():

    GetTime()

    #open URL for zoom meeting
    print('Open Zoom meeting',hour,':', minute,)
    webb.open_new(zoom_link.value) 
    t.sleep(5)

    cords = pe.locateCenterOnScreen("Media//Zoom//LunchMeeting.png", confidence=0.8)
    if cords != 0:
        pe.click(cords)
        t.sleep(10)
    
    #geting the page full size
    with pe.hold('alt'):
        pe.press('space')
        
    t.sleep(1)    
    pe.press('down', presses=4)
    t.sleep(1)
    pe.press('enter')
    t.sleep(15) 
#-----------------------------#
def shered_scr_zoom():  #run the check share screen zoom

    GetTime()
    print('Cheking the screen',hour,':', minute,':', sec)
    
    #first get rid of the fuul screen share
    cords = pe.locateCenterOnScreen("Media//Zoom//Shared3Zoom.jpg", confidence=0.8)
    if cords != 0: #generated only, if it found on the screen
        print('Found Shared3Zoom')
        pe.click(cords)
        t.sleep(1)
        pe.press('esc')
        cords = 0

    #lets find the local host - lecturer
    cords = pe.locateCenterOnScreen("Media//Zoom//Shared1Zoom.jpg", confidence=0.65)
    if cords != 0:
        print('Found Shared1Zoom')
        pe.click(cords)
        t.sleep(1)
        cords = 0

    #click on him and get back to the main screen
    cords = pe.locateCenterOnScreen("Media//Zoom//Shared2Zoom.png", confidence=0.8)
    if cords != 0:
        print('Found Shared2Zoom')
        pe.click(cords)
        t.sleep(1)
        cords = pe.locateCenterOnScreen("Media//Zoom//SharedPinZoom.jpg", confidence=0.8)
        pe.click(cords)
        t.sleep(1)
        cords = 0
#-----------------------------#
def write_name():
    #lets write the course name
    pc.copy(course_name_1.value) #copy the value into clipboard

    with pe.hold('ctrl'):
        pe.press('v')

    pe.write(' - ')

    #lets write the lecturer name
    pc.copy(lecturer_name.value) #copy the value into clipboard

    with pe.hold('ctrl'):
        pe.press('v')

    pe.write(' - ')

    #lets wriite the date
    pe.write(str(month_day))
    pe.write('.')
    pe.write(str(month))
    pe.write('.')
    pe.write(str(year))   
#-----------------------------#
def upload_youtube1():

    global i

    #open YouTube
    print('Open YouTube')
    webb.open_new('www.youtube.com')
    t.sleep(10)    
    
    #lets go upload the file
    cords = pe.locateCenterOnScreen("Media//YouTube//Upload1YouTube.jpg", confidence=0.8)
    pe.click(cords)
    t.sleep(5)

    cords = pe.locateCenterOnScreen("Media//YouTube//Upload2YouTube.jpg", confidence=0.8)
    pe.click(cords)
    t.sleep(5)

    cords = pe.locateCenterOnScreen("Media//YouTube//Upload3YouTube.jpg", confidence=0.8)
    pe.click(cords)
    t.sleep(5)
    
    cords = pe.locateCenterOnScreen("Media//YouTube//Desk.jpg", confidence=0.8)
    pe.click(cords)
    t.sleep(5)

    cords = pe.locateCenterOnScreen("Media//YouTube//Upload4YouTube.jpg", confidence=0.8)
    pe.click(cords)
    t.sleep(5)

    pe.write('Sc', interval=0.25 )
    t.sleep(1)
    pe.press('down', presses = i)
    t.sleep(1)
    pe.press('enter')

    GetTime()
    print('File Found',hour,':', minute,':', sec)
    t.sleep(10)
#-----------------------------#
def upload_youtube2():
    cords = pe.locateCenterOnScreen("Media//YouTube//Upload5YouTube.jpg", confidence=0.8)
    pe.click(cords)
    t.sleep(5)

    cords = pe.locateCenterOnScreen("Media//YouTube//Upload5YouTube.jpg", confidence=0.8)
    pe.click(cords)
    t.sleep(5)

    cords = pe.locateCenterOnScreen("Media//YouTube//Upload5YouTube.jpg", confidence=0.8)
    pe.click(cords)
    t.sleep(5)

    cords = pe.locateCenterOnScreen("Media//YouTube//Upload6YouTube.png", confidence=0.9)
    pe.click(cords)
    t.sleep(5)

    cords = pe.locateCenterOnScreen("Media//YouTube//Upload7YouTube.jpg", confidence=0.8)
    pe.click(cords)
    t.sleep(10)

    GetTime()
    print('uplouding to youtube',hour,':', minute,':', sec)
    t.sleep(1)
#--------------------------------#
def Change_To_English():

    t.sleep(1)

    cords = pe.locateCenterOnScreen("Media//Hebrew.png", confidence=0.9)
    if cords != 0:
        print('Found Hebrew Languge')
        pe.click(cords)
        t.sleep(5)

    cords = pe.locateCenterOnScreen("Media//English.png", confidence=0.9)
    if cords != 0:
        print('Changing to Enlgish')
        pe.click(cords)
        t.sleep(5) 

#=====================================================#
#===================MAIN==CODE========================#
#=====================================================#

wb = load_workbook('LIST.xlsx') #loading new workbook
ws = wb.active #making the worksheet active to work on

GetTime()
Change_To_English()

while(day != 4 or day != 3):

    GetTime()
    
    if(day == 6): #Sunday
        
        if (hour == 9 and minute == 55):

            course_name_1 = ws.cell(row=4, column=2)
            GetXlVal()

            zoom()
            Filmora()  
            StartRecFilmora()
            
            while(hour < 13):
                t.sleep(300)
                GetTime()
                #shered_scr_zoom()
                
            if (hour == 13):
                end_record()
                close_Filmora()

                i = 1
                upload_youtube1()
                write_name()
                upload_youtube2()
#=====================================================#            
    if(day == 0): #monday
        
        if (hour == 7 and minute == 55): #course 1

            course_name_1 = ws.cell(row=2, column=3)
            GetXlVal()

            zoom()
            Filmora()
            StartRecFilmora()
            
            while(hour < 12):
                t.sleep(300)
                GetTime()
                #shered_scr_zoom()
                
            if (hour == 12): #mayeb TAB it one time
                end_record()
                close_Filmora()

                i = 2
                upload_youtube1()
                write_name()
                upload_youtube2()
            
        if (hour == 12 and minute == 55): #course 2

            course_name_1 = ws.cell(row=7, column=3)
            GetXlVal()

            zoom()
            Filmora()
            StartRecFilmora()
            
            while(hour <15 and minute < 55):
                t.sleep(300)
                GetTime() 
                #shered_scr_zoom()
                    
            if (hour == 15 and minute == 55):
                end_record()
                close_Filmora()


                i = 3
                upload_youtube1()
                write_name()
                upload_youtube2()

        if (hour == 17): #course 3

            course_name_1 = ws.cell(row=10, column=3)
            GetXlVal()

            zoom()
            Filmora()
            StartRecFilmora()
            
            while(hour < 18):
                t.sleep(300)
                GetTime()
                #shered_scr_zoom()
                
            if (hour == 18):
                end_record()
                close_Filmora()

                i = 3
                upload_youtube1()
                write_name()
                upload_youtube2()
#=====================================================#            
    if(day == 1): #Thusday
        
        if (hour == 13 and minute == 55):

            course_name_1 = ws.cell(row=8, column=4)
            GetXlVal()

            zoom()
            Filmora()
            StartRecFilmora()
            
            while(hour <17):
                t.sleep(300)
                GetTime()
                #shered_scr_zoom()
                
            if (hour == 17):
                end_record()
                close_Filmora()

                i = 5
                upload_youtube1()
                write_name()
                upload_youtube2()

#=====================================================#             
    if(day == 2): #Wensday
        
        if (hour == 16 and minute == 55):

            course_name_1 = ws.cell(row=11, column=5)
            GetXlVal()

            zoom()
            Filmora()
            StartRecFilmora()
            
            while(hour <20):
                t.sleep(300)
                GetTime()
                #shered_scr_zoom()
                
            if (hour == 20):
                end_record()
                close_Filmora()

                i = 6
                upload_youtube1()
                write_name()
                upload_youtube2()
            
